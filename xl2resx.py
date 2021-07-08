from collections import OrderedDict
import subprocess
import os
from shutil import copyfile
from openpyxl import load_workbook 


def convert_xlsx_to_resx(filename):
    for ws in load_workbook(filename):
        res_dict = OrderedDict({c[0]: c[1:] for c in ws.iter_cols(values_only=True)})
        keys = list(res_dict.keys())

        for k in keys[1:]:
            with open('tmp.txt', 'w') as f:
                f.writelines([f'{n}={v}\n' for n, v in zip(res_dict['name'], res_dict[k])])

            subprocess.run(['resgen', 'tmp.txt', f'/root/workspace/output_/{ws.title}.{k}.resx'])

        copyfile(f'/root/workspace/output_/{ws.title}.{keys[1]}.resx', f'/root/workspace/output_/{ws.title}.resx')

    os.remove('tmp.txt')


if __name__ == '__main__':
    convert_xlsx_to_resx('test.xlsx')
